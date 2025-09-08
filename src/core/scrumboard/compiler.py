import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


class ScrumboardExcelExporter:
    def __init__(self, scrumboard):
        """
        Initialiseer de exporter.
        :param scrumboard: dict van kolommen met lijst taken
        """
        self.scrumboard = scrumboard

    def save_to_bytes(self):
        """Genereer Excel bestand volgens voorbeeldstructuur."""
        print(self.scrumboard)
        wb = Workbook()
        ws = wb.active
        ws.title = "Scrumboard"

        # Kleuren/styling
        header_fill = PatternFill(start_color="4B6CB7", end_color="4B6CB7", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        subheader_fill = PatternFill(start_color="9DC8E8", end_color="9DC8E8", fill_type="solid")
        subheader_font = Font(bold=True)

        # Kolommen per boardkolom
        columns = list(self.scrumboard.keys())
        subheaders = ["Taak", "Prio", "Duur in minuten"]

        # Eerste rij (merged headers)
        col_index = 1
        for col in columns:
            ws.merge_cells(
                start_row=1,
                start_column=col_index,
                end_row=1,
                end_column=col_index + 2
            )
            cell = ws.cell(row=1, column=col_index, value=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_index += 3

        # Tweede rij (subheaders)
        col_index = 1
        for col in columns:
            for sub in subheaders:
                cell = ws.cell(row=2, column=col_index, value=sub)
                cell.fill = subheader_fill
                cell.font = subheader_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_index += 1

        # Data (taken)
        max_len = max(len(tasks) for tasks in self.scrumboard.values())
        for i in range(max_len):
            col_index = 1
            for col, tasks in self.scrumboard.items():
                if i < len(tasks):
                    task = tasks[i]
                    ws.cell(row=i+3, column=col_index, value=task["title"] + "\n" + task["content"]).alignment = Alignment(wrap_text=True, vertical="top")
                    ws.cell(row=i+3, column=col_index+1, value=task["priority"])
                    ws.cell(row=i+3, column=col_index+2, value=task["time_estimate"])
                col_index += 3

        # Kolombreedtes aanpassen
        for i in range(1, col_index):
            ws.column_dimensions[get_column_letter(i)].width = 25

        # Output naar bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def run(self):
        return self.save_to_bytes()
